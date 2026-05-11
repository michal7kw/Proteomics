# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Produce Metascape-ready CSV and XLSX files from the _clean lists.

For each isoform-handling folder (isoforms_distinct, isoforms_collapsed) the
script writes a new 'metascape/' subfolder containing:

  Per-list files (for single-list uploads)
    Ex15_common_accession.csv / .xlsx         header: Accession
    Ex15_common_gene.csv      / .xlsx         header: Gene
    Ex15_POS_specific_accession.csv / .xlsx   ...
    Ex15_POS_specific_gene.csv      / .xlsx
    Ex15_NEG_specific_accession.csv / .xlsx
    Ex15_NEG_specific_gene.csv      / .xlsx

  Combined "Multiple Gene Lists" files (for comparative Express Analysis)
    Ex15_MultipleLists_accession.csv / .xlsx  headers: Common,POS_specific,NEG_specific
    Ex15_MultipleLists_gene.csv      / .xlsx

All files are plain UTF-8 (no BOM), whitespace-stripped, with a single header
row and no blank lines inside a column's data block.
"""

from __future__ import annotations

import csv
from itertools import zip_longest
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).parent
FOLDERS = ("isoforms_distinct", "isoforms_collapsed")

# (short label used as a column header in the combined file, source filename stem)
LISTS = (
    ("Common",        "Ex15_common"),
    ("POS_specific",  "Ex15_POS_specific"),
    ("NEG_specific",  "Ex15_NEG_specific"),
)

# (suffix of the source file, header name, descriptive tag used in outputs)
ID_TYPES = (
    ("_clean.txt",       "Accession", "accession"),
    ("_clean_genes.txt", "Gene",      "gene"),
)


def read_ids(path: Path) -> list[str]:
    """Read a _clean file, returning a whitespace-stripped, blank-free list."""
    out: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s:
            out.append(s)
    return out


def write_csv(path: Path, header: list[str], rows: list[list[str]]) -> None:
    """Write UTF-8 CSV *without* BOM, LF line endings, CRLF inside fields stripped."""
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(header)
        for row in rows:
            w.writerow(row)


def write_xlsx_single(path: Path, header: str, values: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = header
    ws.cell(row=1, column=1, value=header)
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=1, value=v)
    ws.column_dimensions["A"].width = max(12, min(24, max((len(v) for v in values), default=12) + 2))
    wb.save(path)


def write_xlsx_multi(path: Path, headers: list[str], columns: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MultipleGeneLists"
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    for col_idx, values in enumerate(columns, start=1):
        for row_idx, v in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=v)
        letter = ws.cell(row=1, column=col_idx).column_letter
        width = max(12, min(24, max((len(v) for v in values), default=12) + 2))
        ws.column_dimensions[letter].width = width
    wb.save(path)


def process_folder(folder: Path) -> None:
    out_dir = folder / "metascape"
    out_dir.mkdir(exist_ok=True)
    print(f"\n{folder.name}/ -> {out_dir.name}/")

    for src_suffix, header_name, tag in ID_TYPES:
        # --- individual per-list files ----------------------------------
        per_list_values: dict[str, list[str]] = {}
        for col_label, stem in LISTS:
            src = folder / f"{stem}{src_suffix}"
            if not src.exists():
                print(f"  (missing) {src.name} - skipped")
                continue
            ids = read_ids(src)
            per_list_values[col_label] = ids

            csv_path = out_dir / f"{stem}_{tag}.csv"
            xlsx_path = out_dir / f"{stem}_{tag}.xlsx"
            write_csv(csv_path, [header_name], [[v] for v in ids])
            write_xlsx_single(xlsx_path, header_name, ids)
            print(f"  {csv_path.name}  ({len(ids)} {tag}s)")
            print(f"  {xlsx_path.name}")

        # --- combined Multiple Gene Lists file --------------------------
        if per_list_values:
            headers = [lbl for lbl, _ in LISTS if lbl in per_list_values]
            columns = [per_list_values[lbl] for lbl in headers]
            # Pad with blanks so all columns are the same length
            padded = list(zip_longest(*columns, fillvalue=""))
            combined_csv = out_dir / f"Ex15_MultipleLists_{tag}.csv"
            combined_xlsx = out_dir / f"Ex15_MultipleLists_{tag}.xlsx"
            write_csv(combined_csv, headers, [list(row) for row in padded])
            write_xlsx_multi(combined_xlsx, headers, columns)
            sizes = ", ".join(f"{h}={len(v)}" for h, v in zip(headers, columns))
            print(f"  {combined_csv.name}   [{sizes}]")
            print(f"  {combined_xlsx.name}")


def main() -> None:
    for folder_name in FOLDERS:
        process_folder(HERE / folder_name)
    print("\nDone. Recommended for Metascape Express Analysis:")
    print("  isoforms_distinct/metascape/Ex15_MultipleLists_gene.xlsx")


if __name__ == "__main__":
    main()
